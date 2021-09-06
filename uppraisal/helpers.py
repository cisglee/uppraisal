# -*- coding: utf-8 -*-

from copy import deepcopy
import typing as tp
import warnings

import requests
import openpyxl

from .constants import *


def get_excel(
    in_filepath: str,
    worksheet: tp.Optional[tp.Union[str, int]],
    include_fieldnames: tp.Optional[tp.Iterable[str]] = None,
    exclude_fieldnames: tp.Optional[tp.Iterable[str]] = None,
    has_headers: bool = True,
    return_headers: bool = True,
    case_sensitive: bool = False,
    ignore_missing_headers: bool = False
) -> tp.List[tp.List[tp.Any]]:
    """
    Function for reading data from Excel files to a list of lists.

    :param in_filepath: A filepath string referring to an Excel file.
    :type in_filepath: str
    :param worksheet: The name or index number of the worksheet from which data should be retrieved.
    :type worksheet: str or int or None
    :param include_fieldnames: List of fieldnames that should be returned from Excel workbook
    :type include_fieldnames: list
    :param exclude_fieldnames: list of fieldnames that should be ignored when retrieving data from Excel workbook
    :type exclude_fieldnames: list
    :param has_headers: Boolean indicating whether workbook has headers in the first row (True) or not (False)
    :type has_headers: bool
    :param return_headers: Boolean indicating whether the headers should be returned
    :type return_headers: bool
    :param case_sensitive: Boolean indicating whether include_fieldnames or exclude_fieldnames should be sensitive to capitalization
    :type case_sensitive: bool
    :param ignore_missing_headers: Boolean indicating whether missing headers from include_fieldnames or exclude_fieldnames should be ignored (True) or not (False)
    :type ignore_missing_headers: bool
    :return: A list of lists containing the data from the worksheet.
    :rtype: list
    """

    # TODO: Check fieldnames
    # TODO: Check for duplicate headers

    if (include_fieldnames or exclude_fieldnames) and not has_headers:
        raise Exception("Cannot include or exclude data from Excel file by column names when there are no headers.")

    if include_fieldnames and exclude_fieldnames:
        raise Exception("Cannot import data based on inclusion and exclusion simultaneously.")

    if return_headers and not has_headers:
        raise Exception("Cannot return headers from Excel file when there are no headers.")

    if case_sensitive or ignore_missing_headers and not has_headers:
        warnings.warn("Arguments \"capital_sensitive\" and \"ignore_missing_headers\" have no effect when there are no headers.")

    workbook = openpyxl.load_workbook(in_filepath, read_only=True)
    if worksheet is None:
        raise IOError("The name or index number of the worksheet has not been specified.")
    elif isinstance(worksheet, int):
        worksheet = workbook.worksheets[worksheet]
    elif isinstance(worksheet, str):
        worksheet = workbook.get_sheet_by_name(worksheet)

    def strip_case(s: str) -> str:
        return (s.lower() if not case_sensitive else s).strip()

    num_rows: int = worksheet.max_row
    num_cols: int = worksheet.max_column
    curr_row: int = 0
    fieldmap: tp.List[tp.Optional[int]]

    if include_fieldnames:
        org_fieldnames = include_fieldnames
        include_fieldnames = [strip_case(field) for field in include_fieldnames]
        data_template = [None] * len(include_fieldnames)
        fieldmap = [None] * (num_cols + 1)

    elif exclude_fieldnames:
        org_fieldnames = include_fieldnames  # TODO what does this accomplish? include_fieldnames is None here
        exclude_fieldnames = [strip_case(field) for field in exclude_fieldnames]
        data_template = [None] * (num_cols + 1 - len(exclude_fieldnames))
        fieldmap = [None] * (num_cols + 1)

    else:
        data_template = [None] * (num_cols + 1)
        fieldmap = [i for i in range(0, num_cols + 1)]

    data = []
    fieldnames_found = []
    while curr_row < num_rows:
        curr_row += 1

        if curr_row == 1 and has_headers:

            if include_fieldnames:
                curr_col = 0
                while curr_col < num_cols:
                    curr_col += 1
                    cell_value = worksheet.cell(curr_row, curr_col).value
                    if cell_value in include_fieldnames:
                        fieldmap[curr_col] = include_fieldnames.index(cell_value)
                        fieldnames_found.append(cell_value)

            elif exclude_fieldnames:
                include_fieldnames = []
                curr_col = 0
                while curr_col < num_cols:
                    curr_col += 1
                    cell_value = strip_case(worksheet.cell(curr_row, curr_col).value)
                    if cell_value not in exclude_fieldnames:
                        include_fieldnames.append(cell_value)
                        fieldnr = max(fieldmap) + 1 if 0 in fieldmap else 0
                        fieldmap[curr_col] = fieldnr
                        fieldnames_found.append(cell_value)
            else:
                # If not include or exclude fieldnames, include_fieldnames is set to all fieldnames
                org_fieldnames = [c.value for c in worksheet[1]]

            if has_headers and return_headers:
                data.append(org_fieldnames)

        else:
            row_data = deepcopy(data_template)  # TODO No need for deep, template is always flat?
            curr_col = 0
            while curr_col < num_cols:
                curr_col += 1
                fieldnr = fieldmap[curr_col]
                if fieldnr is not None:
                    row_data[fieldnr] = worksheet.cell(curr_row, curr_col).value

            data.append(row_data)

    if ignore_missing_headers is False:
        missing = None
        if include_fieldnames:
            missing = set(include_fieldnames) - set(fieldnames_found)
        elif exclude_fieldnames:
            missing = set(exclude_fieldnames) - set(fieldnames_found)
        if missing:
            raise IOError(f"The following fieldnames could not be found: {', '.join(missing)}")

    return data


def list_submissions(
    cv_access_token: str,
    cv_course_id: int,
    cv_assignment_id: int,
    sort_by: str = DEFAULT_SORT_BY,
    select_columns: tp.Optional[tp.Iterable[str]] = DEFAULT_SELECT_COLUMNS,
    out_filepath: tp.Optional[str] = DEFAULT_OUT_FILEPATH,
) -> tp.List[tp.Dict[str, tp.Any]]:
    """
    Retrieve submission meta data for a Canvas assignment

    :param cv_access_token: Canvas access token (generally 70 characters in length; see
    https://canvas.instructure.com/courses/785215/pages/getting-started-with-the-api).
    :type cv_access_token: str
    :param cv_course_id: Course identifier (see Canvas course URL).
    :type cv_course_id: int
    :param cv_assignment_id: Assignment identifier (see Canvas assignment URL).
    :type cv_assignment_id: int
    :param sort_by: Assignment meta data column by which the return data should be sorted (e.g., 'submitted_at'). Default
    is 'user_sortable_name'.
    :type sort_by: str
    :param select_columns: Iterable with column names to return, empty or None if all meta data should be returned.
    :type select_columns: None or list
    :param out_filepath: If this is a string, save to the given path.
    :type out_filepath: str or None
    :return: List of dictionaries containing meta data for assignments submitted to assignment id `cv_id_assignment`
    :rtype: list
    """

    def filter_assignments(rbuffer):
        assignments = []
        for assignment in rbuffer.json():
            if assignment['submitted_at']:

                if 'user' in assignment:
                    for key in assignment['user'].keys():
                        assignment['user_' + key] = assignment['user'][key]
                    del assignment['user']

                if 'attachments' in assignment:
                    attachments = []
                    for attachment in assignment['attachments']:
                        attachments.append(attachment['filename'])
                    assignment['attachments'] = '; '.join(attachments)

                if select_columns:
                    remove_columns = []
                    for col in assignment:
                        if col not in select_columns:
                            remove_columns.append(col)
                    for col in remove_columns:
                        del assignment[col]

                assignments.append(assignment)

        return assignments

    if sort_by and select_columns and sort_by not in select_columns:
        raise IOError('The column specified for sortby must be in the list of columns you would like to be returned')

    params = {'grouped': True, 'per_page': 100, 'include': 'user'}
    headers = {'Authorization': 'Bearer {}'.format(cv_access_token), 'Content-type': 'application/json'}
    assignments = []

    url = CANVAS_SUBMISSIONS_URL.format(cid=cv_course_id, aid=cv_assignment_id)
    response = requests.get(url, headers=headers, params=params)
    if response.status_code // 100 != 2:
        raise Exception(f'Got an error {response.status_code} from {url}: {response.text}')

    assignments += filter_assignments(response)
    while response.links['current']['url'] != response.links['last']['url']:
        url = response.links['next']['url']
        response = requests.get(url, headers=headers, params=params)
        if response.status_code // 100 != 2:
            raise Exception(f'Got an error {response.status_code} from {url}: {response.text}')

        assignments += filter_assignments(response)

    if sort_by in assignments[0]:
        assignments.sort(key=lambda k: k[sort_by])
    else:
        if type(sort_by) == str:
            warnings.warn(f'The specified sort key (i.e., {sort_by}) was not found')
        else:
            warnings.warn(f'The parameter sortby should be a string, got {sort_by}')

    if select_columns:
        headers = list(select_columns)
    else:
        _headers: tp.Dict[str, None] = dict()
        for assignment in assignments:
            for header in assignment.keys():
                _headers[header] = True
        headers = list(_headers.keys())
        del _headers

    assignment_list = [headers]
    for assignment in assignments:
        assignment_row = []
        for header in headers:
            assignment_row.append(assignment[header])
        assignment_list.append(assignment_row)

    if out_filepath:
        put_excel(out_filepath, assignment_list)

    return assignments


# TODO Not used?
_T = tp.TypeVar('_T')


def put_excel(
    out_filepath: str,
    data: tp.List[tp.List[tp.Any]],
    worksheet_names: tp.Optional[tp.Union[str, tp.List[str]]] = None
) -> None:
    """
    Method for writing data from a list of lists to Excel

    :Examples:
    >>> dat = [["Header 1", "Header 2", "Header 3"], ["row 1 cell 1", "row 1 cell 2", "row 1 cell 3"], ["row 2 cell 1", "row 2 cell 2", "row 3 cell 3"], ["row 3 cell 1", "row 3 cell 2", "row 3 cell 3"]]
    >>> put_excel("./example1.xls", dat) # Example writing to a single worksheet
    >>> dat = [[["Header 1", "Header 2", "Header 3"], ["row 1 cell 1", "row 1 cell 2", "row 1 cell 3"], ["row 2 cell 1", "row 2 cell 2", "row 3 cell 3"], ["row 3 cell 1", "row 3 cell 2", "row 3 cell 3"]], [["Header 4", "Header 5", "Header 6"], ["row 4 cell 1", "row 4 cell 2", "row 4 cell 3"], ["row 5 cell 1", "row 5 cell 2", "row 5 cell 3"], ["row 6 cell 1", "row 6 cell 2", "row 6 cell 3"]], [["Header 7", "Header 8", "Header 9"], ["row 7 cell 1", "row 7 cell 2", "row 7 cell 3"], ["row 8 cell 1", "row 8 cell 2", "row 8 cell 3"], ["row 9 cell 1", "row 9 cell 2", "row 9 cell 3"]]]
    >>> put_excel("./example2.xls", dat, worksheet_names=["WS1", "WS2", "WS3"]) # Example writing to multiple worksheets

    :param data: List of lists containing the data that should be written to Excel.
    :type data: list
    :param out_filepath: File path pointing to location and file name to which the data should be written.
    :type out_filepath: str
    :param worksheet_names: List of names of worksheets to which the data should be written or string name for the
    worksheet.
    :type worksheet_names: list or str
    :return: None
    """

    wb = openpyxl.Workbook()
    l = 3 if isinstance(data[0][0], list) else 2
    if l == 2:
        # If data is list of rows, structure as if it were a list containing lists of rows
        if worksheet_names:
            if isinstance(worksheet_names, str):
                worksheet_names = [worksheet_names]
        data = [data]
        l = 3

    if l == 3:
        # If data is list containing lists of rows create a worksheet for every list of rows
        # Check if there are names for every worksheet
        if worksheet_names:
            if len(data) != len(worksheet_names):
                raise IOError("Data structure indicates {} worksheet, but {} worksheet name{} been provided".format(
                    len(data), len(worksheet_names), "s have" if len(worksheet_names) > 1 else " has"
                ))
        else:
            worksheet_names = [f'Sheet{i + 1}' for i in range(len(data))]

        for index, (wsname, wsdat) in enumerate(zip(worksheet_names, data)):
            ws = wb.create_sheet(wsname, index)
            for row_j, row in enumerate(wsdat, 1):
                for cell_i, cell in enumerate(row, 1):
                    ws.cell(row_j, cell_i, cell)

        wb.save(out_filepath)

    return


def chunker(seq: tp.Sequence[_T], size: int = DEFAULT_CHUNK_SIZE) -> tp.Iterator[tp.Sequence[_T]]:
    """
    Chunks a list of items to help reduce the API load

    :source: https://stackoverflow.com/a/434328

    :param seq: array over which to loop
    :rtype seq: list
    :param size: size of chunks
    :rtype size: int
    :return: chunked list
    :rtype: list
    """

    for pos in range(0, len(seq), size):
        yield seq[pos:pos + size]
